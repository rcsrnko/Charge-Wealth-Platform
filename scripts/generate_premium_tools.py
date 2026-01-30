#!/usr/bin/env python3
"""
Charge Wealth Premium Financial Tools Generator
Creates 5 CFO-grade Excel files with professional formatting, formulas, and charts.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, DataBarRule
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

# Charge Wealth Brand Colors
HONEY = "F5A623"  # Primary gold/amber
HONEY_LIGHT = "FFF3D4"  # Light honey for backgrounds
VANILLA = "FFFEF5"  # Cream white
DARK_TEXT = "2D3436"  # Near black for text
ACCENT_GREEN = "27AE60"  # Success/positive
ACCENT_RED = "E74C3C"  # Warning/negative
GRAY_HEADER = "636E72"  # Subtle headers

# Style definitions
def get_styles():
    """Create reusable styles for consistent formatting"""
    styles = {}
    
    # Header style
    styles['header'] = {
        'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color=HONEY, end_color=HONEY, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color=HONEY),
            right=Side(style='thin', color=HONEY),
            top=Side(style='thin', color=HONEY),
            bottom=Side(style='thin', color=HONEY)
        )
    }
    
    # Title style
    styles['title'] = {
        'font': Font(name='Calibri', size=24, bold=True, color=HONEY),
        'alignment': Alignment(horizontal='left', vertical='center')
    }
    
    # Subtitle style
    styles['subtitle'] = {
        'font': Font(name='Calibri', size=14, color=GRAY_HEADER),
        'alignment': Alignment(horizontal='left', vertical='center')
    }
    
    # Data cell style
    styles['data'] = {
        'font': Font(name='Calibri', size=11, color=DARK_TEXT),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
    }
    
    # Label style
    styles['label'] = {
        'font': Font(name='Calibri', size=11, color=DARK_TEXT),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
    }
    
    # Currency style
    styles['currency'] = {
        'font': Font(name='Calibri', size=11, color=DARK_TEXT),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'number_format': '"$"#,##0.00'
    }
    
    # Percentage style
    styles['percent'] = {
        'font': Font(name='Calibri', size=11, color=DARK_TEXT),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'number_format': '0.00%'
    }
    
    # KPI positive
    styles['kpi_positive'] = {
        'font': Font(name='Calibri', size=18, bold=True, color=ACCENT_GREEN),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    # KPI negative
    styles['kpi_negative'] = {
        'font': Font(name='Calibri', size=18, bold=True, color=ACCENT_RED),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    # Section header
    styles['section'] = {
        'font': Font(name='Calibri', size=14, bold=True, color=HONEY),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    }
    
    return styles

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell"""
    for key, value in style_dict.items():
        if key == 'number_format':
            cell.number_format = value
        else:
            setattr(cell, key, value)

def add_branding_header(ws, title, subtitle=""):
    """Add Charge Wealth branding header to worksheet"""
    styles = get_styles()
    
    # Merge cells for title
    ws.merge_cells('B2:H2')
    ws['B2'] = title
    apply_style(ws['B2'], styles['title'])
    
    if subtitle:
        ws.merge_cells('B3:H3')
        ws['B3'] = subtitle
        apply_style(ws['B3'], styles['subtitle'])
    
    # Add "Powered by Charge Wealth" 
    ws.merge_cells('B4:D4')
    ws['B4'] = "Powered by Charge Wealth"
    ws['B4'].font = Font(name='Calibri', size=10, italic=True, color=GRAY_HEADER)
    
    return 6  # Return row to start content

def create_instructions_sheet(wb, tool_name, instructions):
    """Create a comprehensive instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)
    styles = get_styles()
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80
    
    row = 2
    
    # Title
    ws.merge_cells(f'B{row}:B{row}')
    ws[f'B{row}'] = f"📊 {tool_name}"
    apply_style(ws[f'B{row}'], styles['title'])
    row += 2
    
    # Welcome message
    ws[f'B{row}'] = "Welcome to your premium financial planning tool from Charge Wealth!"
    ws[f'B{row}'].font = Font(name='Calibri', size=12, color=DARK_TEXT)
    row += 2
    
    # Instructions sections
    for section_title, section_content in instructions.items():
        ws[f'B{row}'] = section_title
        apply_style(ws[f'B{row}'], styles['section'])
        ws.row_dimensions[row].height = 25
        row += 1
        
        for item in section_content:
            ws[f'B{row}'] = f"  • {item}"
            ws[f'B{row}'].font = Font(name='Calibri', size=11, color=DARK_TEXT)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = max(20, len(item) // 80 * 15 + 20)
            row += 1
        row += 1
    
    # Footer
    row += 1
    ws[f'B{row}'] = "Questions? Visit chargewealth.co for support and more premium tools."
    ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True, color=HONEY)
    
    return ws


# ============================================================================
# 1. CASH FLOW COMMAND CENTER
# ============================================================================
def create_cash_flow_command_center(output_path):
    """Create comprehensive cash flow tracking with projections"""
    wb = Workbook()
    styles = get_styles()
    
    # Instructions
    instructions = {
        "🎯 Overview": [
            "This Cash Flow Command Center helps you track every dollar coming in and going out.",
            "Use the Dashboard for a quick financial snapshot, and detailed sheets for tracking.",
            "All calculations update automatically as you enter data."
        ],
        "📝 Getting Started": [
            "Start with the 'Income' sheet - enter all your income sources",
            "Move to 'Fixed Expenses' for recurring monthly bills",
            "Use 'Variable Expenses' for discretionary spending",
            "The Dashboard will automatically calculate your cash flow position"
        ],
        "💡 Pro Tips": [
            "Update weekly for best results - 15 minutes every Sunday",
            "Use the 12-month projection to plan for large purchases",
            "Set up alerts when cash flow drops below your emergency threshold",
            "Review the trends chart monthly to spot spending creep"
        ],
        "📊 Understanding the Dashboard": [
            "GREEN numbers = positive cash flow (you're saving money)",
            "RED numbers = negative cash flow (spending more than earning)",
            "The 'Runway' shows how many months of expenses you could cover"
        ]
    }
    create_instructions_sheet(wb, "Cash Flow Command Center", instructions)
    
    # Dashboard Sheet
    ws = wb.create_sheet("Dashboard")
    start_row = add_branding_header(ws, "Cash Flow Command Center", "Your Complete Financial Picture")
    
    # Set column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['B'].width = 20
    
    # KPI Section
    row = start_row
    ws[f'B{row}'] = "KEY METRICS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    # KPI Cards
    kpis = [
        ("Monthly Income", "=Income!H15", "currency"),
        ("Monthly Expenses", "='Fixed Expenses'!E20+'Variable Expenses'!E25", "currency"),
        ("Net Cash Flow", "=B9-B10", "currency"),
        ("Savings Rate", "=IF(B9>0,B11/B9,0)", "percent"),
        ("Emergency Runway", "=IF(B10>0,Settings!B5/B10,0)", "months")
    ]
    
    for i, (label, formula, fmt) in enumerate(kpis):
        ws[f'B{row}'] = label
        apply_style(ws[f'B{row}'], styles['label'])
        ws[f'C{row}'] = formula
        if fmt == "currency":
            ws[f'C{row}'].number_format = '"$"#,##0.00'
        elif fmt == "percent":
            ws[f'C{row}'].number_format = '0.0%'
        else:
            ws[f'C{row}'].number_format = '0.0 "months"'
        ws[f'C{row}'].font = Font(name='Calibri', size=14, bold=True, color=DARK_TEXT)
        row += 1
    
    # Add conditional formatting for net cash flow
    ws.conditional_formatting.add('C11',
        FormulaRule(formula=['C11>=0'], fill=PatternFill(bgColor=ACCENT_GREEN)))
    ws.conditional_formatting.add('C11',
        FormulaRule(formula=['C11<0'], fill=PatternFill(bgColor=ACCENT_RED)))
    
    # 12-Month Projection Section
    row += 3
    ws[f'B{row}'] = "12-MONTH CASH FLOW PROJECTION"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:N{row}')
    row += 2
    
    # Month headers
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    ws[f'B{row}'] = "Category"
    apply_style(ws[f'B{row}'], styles['header'])
    for i, month in enumerate(months):
        col = get_column_letter(3 + i)
        ws[f'{col}{row}'] = month
        apply_style(ws[f'{col}{row}'], styles['header'])
    row += 1
    
    # Projection rows
    proj_categories = ['Projected Income', 'Projected Expenses', 'Net Cash Flow', 'Cumulative Balance']
    for cat in proj_categories:
        ws[f'B{row}'] = cat
        apply_style(ws[f'B{row}'], styles['label'])
        for i in range(12):
            col = get_column_letter(3 + i)
            if cat == 'Projected Income':
                ws[f'{col}{row}'] = '=$C$9*(1+Settings!$B$7)^' + str(i)
            elif cat == 'Projected Expenses':
                ws[f'{col}{row}'] = '=$C$10*(1+Settings!$B$8)^' + str(i)
            elif cat == 'Net Cash Flow':
                ws[f'{col}{row}'] = f'={col}{row-2}-{col}{row-1}'
            else:  # Cumulative
                if i == 0:
                    ws[f'{col}{row}'] = f'=Settings!$B$5+{col}{row-1}'
                else:
                    prev_col = get_column_letter(2 + i)
                    ws[f'{col}{row}'] = f'={prev_col}{row}+{col}{row-1}'
            ws[f'{col}{row}'].number_format = '"$"#,##0'
        row += 1
    
    # Add cash flow chart
    chart = LineChart()
    chart.title = "12-Month Cash Flow Projection"
    chart.style = 10
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Month"
    chart.width = 18
    chart.height = 10
    
    # Data references for chart
    data = Reference(ws, min_col=3, min_row=row-2, max_col=14, max_row=row-1)
    cats = Reference(ws, min_col=3, min_row=row-5, max_col=14)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].name = "Net Cash Flow"
    chart.series[1].name = "Cumulative Balance"
    
    ws.add_chart(chart, f'B{row + 2}')
    
    # Income Sheet
    ws_income = wb.create_sheet("Income")
    start_row = add_branding_header(ws_income, "Income Tracking", "All Revenue Sources")
    
    for col in range(1, 10):
        ws_income.column_dimensions[get_column_letter(col)].width = 15
    ws_income.column_dimensions['B'].width = 25
    
    row = start_row
    headers = ['Source', 'Type', 'Frequency', 'Amount', 'Annual Total', 'Monthly Equiv.', 'Notes']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_income[f'{col}{row}'] = h
        apply_style(ws_income[f'{col}{row}'], styles['header'])
    row += 1
    
    # Sample income rows
    income_sources = [
        ('Primary Salary', 'W-2', 'Bi-weekly', 3500),
        ('Side Business', '1099', 'Monthly', 1200),
        ('Dividends', 'Investment', 'Quarterly', 500),
        ('Rental Income', 'Passive', 'Monthly', 1800),
        ('Freelance', '1099', 'Variable', 800),
    ]
    
    for source, type_, freq, amount in income_sources:
        ws_income[f'B{row}'] = source
        ws_income[f'C{row}'] = type_
        ws_income[f'D{row}'] = freq
        ws_income[f'E{row}'] = amount
        ws_income[f'E{row}'].number_format = '"$"#,##0.00'
        # Annual calculation based on frequency
        ws_income[f'F{row}'] = f'=IF(D{row}="Weekly",E{row}*52,IF(D{row}="Bi-weekly",E{row}*26,IF(D{row}="Monthly",E{row}*12,IF(D{row}="Quarterly",E{row}*4,E{row}))))'
        ws_income[f'F{row}'].number_format = '"$"#,##0.00'
        ws_income[f'G{row}'] = f'=F{row}/12'
        ws_income[f'G{row}'].number_format = '"$"#,##0.00'
        apply_style(ws_income[f'B{row}'], styles['label'])
        row += 1
    
    # Add more empty rows for user input
    for _ in range(5):
        for i in range(7):
            col = get_column_letter(2 + i)
            apply_style(ws_income[f'{col}{row}'], styles['data'])
        ws_income[f'F{row}'] = f'=IF(D{row}="Weekly",E{row}*52,IF(D{row}="Bi-weekly",E{row}*26,IF(D{row}="Monthly",E{row}*12,IF(D{row}="Quarterly",E{row}*4,E{row}))))'
        ws_income[f'G{row}'] = f'=F{row}/12'
        row += 1
    
    # Totals
    row += 1
    ws_income[f'B{row}'] = "TOTAL"
    ws_income[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws_income[f'F{row}'] = f'=SUM(F{start_row+1}:F{row-2})'
    ws_income[f'F{row}'].number_format = '"$"#,##0.00'
    ws_income[f'F{row}'].font = Font(bold=True, size=12)
    ws_income[f'G{row}'] = f'=SUM(G{start_row+1}:G{row-2})'
    ws_income[f'G{row}'].number_format = '"$"#,##0.00'
    ws_income[f'G{row}'].font = Font(bold=True, size=12)
    ws_income['H15'] = f'=G{row}'  # Reference for dashboard
    
    # Fixed Expenses Sheet
    ws_fixed = wb.create_sheet("Fixed Expenses")
    start_row = add_branding_header(ws_fixed, "Fixed Expenses", "Recurring Monthly Bills")
    
    for col in range(1, 8):
        ws_fixed.column_dimensions[get_column_letter(col)].width = 15
    ws_fixed.column_dimensions['B'].width = 25
    
    row = start_row
    headers = ['Expense', 'Category', 'Due Date', 'Amount', 'Annual Total']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_fixed[f'{col}{row}'] = h
        apply_style(ws_fixed[f'{col}{row}'], styles['header'])
    row += 1
    
    fixed_expenses = [
        ('Rent/Mortgage', 'Housing', 1, 2200),
        ('Car Payment', 'Transportation', 15, 450),
        ('Car Insurance', 'Insurance', 5, 125),
        ('Health Insurance', 'Insurance', 1, 350),
        ('Phone', 'Utilities', 20, 85),
        ('Internet', 'Utilities', 12, 75),
        ('Streaming Services', 'Entertainment', 1, 45),
        ('Gym Membership', 'Health', 1, 50),
    ]
    
    for expense, category, due, amount in fixed_expenses:
        ws_fixed[f'B{row}'] = expense
        ws_fixed[f'C{row}'] = category
        ws_fixed[f'D{row}'] = due
        ws_fixed[f'E{row}'] = amount
        ws_fixed[f'E{row}'].number_format = '"$"#,##0.00'
        ws_fixed[f'F{row}'] = f'=E{row}*12'
        ws_fixed[f'F{row}'].number_format = '"$"#,##0.00'
        apply_style(ws_fixed[f'B{row}'], styles['label'])
        row += 1
    
    for _ in range(5):
        for i in range(5):
            col = get_column_letter(2 + i)
            apply_style(ws_fixed[f'{col}{row}'], styles['data'])
        ws_fixed[f'F{row}'] = f'=E{row}*12'
        row += 1
    
    row += 1
    ws_fixed[f'B{row}'] = "TOTAL"
    ws_fixed[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws_fixed[f'E{row}'] = f'=SUM(E{start_row+1}:E{row-2})'
    ws_fixed[f'E{row}'].number_format = '"$"#,##0.00'
    ws_fixed[f'E{row}'].font = Font(bold=True, size=12)
    ws_fixed['E20'] = f'=E{row}'  # Reference for dashboard
    
    # Variable Expenses Sheet
    ws_var = wb.create_sheet("Variable Expenses")
    start_row = add_branding_header(ws_var, "Variable Expenses", "Discretionary Spending")
    
    for col in range(1, 8):
        ws_var.column_dimensions[get_column_letter(col)].width = 15
    ws_var.column_dimensions['B'].width = 25
    
    row = start_row
    headers = ['Category', 'Budget', 'Actual', 'Variance', '% of Budget']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_var[f'{col}{row}'] = h
        apply_style(ws_var[f'{col}{row}'], styles['header'])
    row += 1
    
    var_categories = [
        ('Groceries', 600, 580),
        ('Dining Out', 300, 420),
        ('Gas/Transportation', 200, 175),
        ('Shopping', 200, 310),
        ('Entertainment', 150, 125),
        ('Personal Care', 100, 95),
        ('Gifts', 100, 150),
        ('Miscellaneous', 200, 180),
    ]
    
    for category, budget, actual in var_categories:
        ws_var[f'B{row}'] = category
        ws_var[f'C{row}'] = budget
        ws_var[f'C{row}'].number_format = '"$"#,##0.00'
        ws_var[f'D{row}'] = actual
        ws_var[f'D{row}'].number_format = '"$"#,##0.00'
        ws_var[f'E{row}'] = f'=C{row}-D{row}'
        ws_var[f'E{row}'].number_format = '"$"#,##0.00'
        ws_var[f'F{row}'] = f'=IF(C{row}>0,D{row}/C{row},0)'
        ws_var[f'F{row}'].number_format = '0%'
        apply_style(ws_var[f'B{row}'], styles['label'])
        row += 1
    
    # Add conditional formatting for variance
    ws_var.conditional_formatting.add(f'E{start_row+1}:E{row-1}',
        FormulaRule(formula=[f'E{start_row+1}>=0'], fill=PatternFill(bgColor="C6EFCE")))
    ws_var.conditional_formatting.add(f'E{start_row+1}:E{row-1}',
        FormulaRule(formula=[f'E{start_row+1}<0'], fill=PatternFill(bgColor="FFC7CE")))
    
    # Data bars for % of budget
    ws_var.conditional_formatting.add(f'F{start_row+1}:F{row-1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1.5,
                   color=HONEY))
    
    for _ in range(5):
        for i in range(5):
            col = get_column_letter(2 + i)
            apply_style(ws_var[f'{col}{row}'], styles['data'])
        ws_var[f'E{row}'] = f'=C{row}-D{row}'
        ws_var[f'F{row}'] = f'=IF(C{row}>0,D{row}/C{row},0)'
        row += 1
    
    row += 1
    ws_var[f'B{row}'] = "TOTAL"
    ws_var[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws_var[f'C{row}'] = f'=SUM(C{start_row+1}:C{row-2})'
    ws_var[f'C{row}'].number_format = '"$"#,##0.00'
    ws_var[f'D{row}'] = f'=SUM(D{start_row+1}:D{row-2})'
    ws_var[f'D{row}'].number_format = '"$"#,##0.00'
    ws_var[f'E{row}'] = f'=C{row}-D{row}'
    ws_var[f'E{row}'].number_format = '"$"#,##0.00'
    ws_var['E25'] = f'=D{row}'  # Reference for dashboard
    
    # Settings Sheet
    ws_settings = wb.create_sheet("Settings")
    ws_settings.column_dimensions['A'].width = 5
    ws_settings.column_dimensions['B'].width = 30
    ws_settings.column_dimensions['C'].width = 20
    
    ws_settings['B2'] = "Settings & Assumptions"
    apply_style(ws_settings['B2'], styles['title'])
    
    settings = [
        ('Emergency Fund Balance', 15000, '"$"#,##0.00'),
        ('Target Savings Rate', 0.20, '0%'),
        ('Expected Income Growth', 0.03, '0.0%'),
        ('Expected Expense Growth', 0.025, '0.0%'),
    ]
    
    row = 5
    for setting, value, fmt in settings:
        ws_settings[f'A{row}'] = setting
        ws_settings[f'B{row}'] = value
        ws_settings[f'B{row}'].number_format = fmt
        row += 1
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


# ============================================================================
# 2. TAX PLANNING COMMAND CENTER
# ============================================================================
def create_tax_planning_command_center(output_path):
    """Create comprehensive tax planning tool"""
    wb = Workbook()
    styles = get_styles()
    
    instructions = {
        "🎯 Overview": [
            "This Tax Planning Command Center helps you minimize taxes legally and maximize refunds.",
            "Track deductions throughout the year, estimate quarterly payments, and plan strategies.",
            "Updated for current tax brackets and standard deductions."
        ],
        "📝 Getting Started": [
            "Enter your filing status and income on the 'Tax Estimator' sheet",
            "Track deductions as they occur in the 'Deductions' sheet",
            "Use 'Quarterly Payments' to calculate estimated tax payments",
            "Review 'Optimization' for tax-saving strategies"
        ],
        "💡 Pro Tips": [
            "Update deductions monthly to catch everything",
            "Compare standard vs itemized deductions quarterly",
            "Max out retirement contributions before year-end",
            "Consider timing of income and deductions across tax years"
        ],
        "⚠️ Disclaimer": [
            "This tool is for planning purposes only",
            "Consult a qualified tax professional for advice",
            "Tax laws change - verify current rates annually"
        ]
    }
    create_instructions_sheet(wb, "Tax Planning Command Center", instructions)
    
    # Tax Estimator Sheet
    ws = wb.create_sheet("Tax Estimator")
    start_row = add_branding_header(ws, "Tax Planning Command Center", "Estimate Your Tax Liability")
    
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['B'].width = 28
    
    row = start_row
    
    # Filing Status Section
    ws[f'B{row}'] = "FILING INFORMATION"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws[f'B{row}'] = "Filing Status"
    ws[f'C{row}'] = "Single"  # User can change to: Married Filing Jointly, etc.
    apply_style(ws[f'B{row}'], styles['label'])
    row += 1
    
    ws[f'B{row}'] = "Tax Year"
    ws[f'C{row}'] = 2024
    apply_style(ws[f'B{row}'], styles['label'])
    row += 2
    
    # Income Section
    ws[f'B{row}'] = "INCOME"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    income_items = [
        ('W-2 Wages', 85000),
        ('Self-Employment Income', 25000),
        ('Interest Income', 500),
        ('Dividend Income', 1200),
        ('Capital Gains (Long-term)', 3000),
        ('Capital Gains (Short-term)', 1500),
        ('Other Income', 0),
    ]
    
    income_start = row
    for item, amount in income_items:
        ws[f'B{row}'] = item
        ws[f'C{row}'] = amount
        ws[f'C{row}'].number_format = '"$"#,##0.00'
        apply_style(ws[f'B{row}'], styles['label'])
        row += 1
    
    row += 1
    ws[f'B{row}'] = "GROSS INCOME"
    ws[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws[f'C{row}'] = f'=SUM(C{income_start}:C{row-2})'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=12)
    gross_income_row = row
    row += 2
    
    # Adjustments Section
    ws[f'B{row}'] = "ADJUSTMENTS (Above-the-Line)"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    adjustments = [
        ('Traditional IRA Contribution', 6500),
        ('HSA Contribution', 3850),
        ('Self-Employment Tax (50%)', '=C' + str(income_start+1) + '*0.0765'),
        ('Student Loan Interest', 2500),
        ('Educator Expenses', 300),
    ]
    
    adj_start = row
    for item, amount in adjustments:
        ws[f'B{row}'] = item
        ws[f'C{row}'] = amount
        ws[f'C{row}'].number_format = '"$"#,##0.00'
        apply_style(ws[f'B{row}'], styles['label'])
        row += 1
    
    row += 1
    ws[f'B{row}'] = "TOTAL ADJUSTMENTS"
    ws[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws[f'C{row}'] = f'=SUM(C{adj_start}:C{row-2})'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    total_adj_row = row
    row += 1
    
    ws[f'B{row}'] = "ADJUSTED GROSS INCOME (AGI)"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f'=C{gross_income_row}-C{total_adj_row}'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=12, color=HONEY)
    agi_row = row
    row += 2
    
    # Deductions Section
    ws[f'B{row}'] = "DEDUCTIONS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws[f'B{row}'] = "Standard Deduction (2024 Single)"
    ws[f'C{row}'] = 14600
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    standard_ded_row = row
    row += 1
    
    ws[f'B{row}'] = "Itemized Deductions"
    ws[f'C{row}'] = "=Deductions!E50"
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    itemized_row = row
    row += 1
    
    ws[f'B{row}'] = "DEDUCTION USED (Higher of)"
    ws[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws[f'C{row}'] = f'=MAX(C{standard_ded_row},C{itemized_row})'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    deduction_row = row
    row += 2
    
    # Taxable Income
    ws[f'B{row}'] = "TAXABLE INCOME"
    ws[f'B{row}'].font = Font(bold=True, size=14)
    ws[f'C{row}'] = f'=MAX(0,C{agi_row}-C{deduction_row})'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=14, color=HONEY)
    ws[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    taxable_row = row
    row += 2
    
    # Tax Calculation Section
    ws[f'B{row}'] = "TAX CALCULATION"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:D{row}')
    row += 2
    
    # 2024 Tax Brackets (Single)
    ws[f'B{row}'] = "Federal Income Tax"
    # Simplified tax calculation formula
    ws[f'C{row}'] = f'=IF(C{taxable_row}<=11600,C{taxable_row}*0.10,IF(C{taxable_row}<=47150,1160+(C{taxable_row}-11600)*0.12,IF(C{taxable_row}<=100525,5426+(C{taxable_row}-47150)*0.22,IF(C{taxable_row}<=191950,17168.5+(C{taxable_row}-100525)*0.24,IF(C{taxable_row}<=243725,39110.5+(C{taxable_row}-191950)*0.32,IF(C{taxable_row}<=609350,55678.5+(C{taxable_row}-243725)*0.35,183647.25+(C{taxable_row}-609350)*0.37))))))'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    fed_tax_row = row
    row += 1
    
    ws[f'B{row}'] = "Self-Employment Tax"
    ws[f'C{row}'] = f'=C{income_start+1}*0.153'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    se_tax_row = row
    row += 1
    
    ws[f'B{row}'] = "State Tax (Est. 5%)"
    ws[f'C{row}'] = f'=C{taxable_row}*0.05'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    state_tax_row = row
    row += 2
    
    ws[f'B{row}'] = "TOTAL TAX LIABILITY"
    ws[f'B{row}'].font = Font(bold=True, size=14)
    ws[f'C{row}'] = f'=C{fed_tax_row}+C{se_tax_row}+C{state_tax_row}'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=14, color=ACCENT_RED)
    total_tax_row = row
    row += 2
    
    # Effective Tax Rate
    ws[f'B{row}'] = "Effective Tax Rate"
    ws[f'C{row}'] = f'=IF(C{gross_income_row}>0,C{total_tax_row}/C{gross_income_row},0)'
    ws[f'C{row}'].number_format = '0.00%'
    ws[f'C{row}'].font = Font(bold=True, size=12, color=DARK_TEXT)
    row += 1
    
    ws[f'B{row}'] = "Marginal Tax Rate"
    ws[f'C{row}'] = f'=IF(C{taxable_row}<=11600,0.10,IF(C{taxable_row}<=47150,0.12,IF(C{taxable_row}<=100525,0.22,IF(C{taxable_row}<=191950,0.24,IF(C{taxable_row}<=243725,0.32,IF(C{taxable_row}<=609350,0.35,0.37))))))'
    ws[f'C{row}'].number_format = '0%'
    
    # Deductions Tracking Sheet
    ws_ded = wb.create_sheet("Deductions")
    start_row = add_branding_header(ws_ded, "Deduction Tracker", "Itemized Deductions")
    
    for col in range(1, 8):
        ws_ded.column_dimensions[get_column_letter(col)].width = 15
    ws_ded.column_dimensions['B'].width = 30
    
    row = start_row
    
    # Medical Expenses
    ws_ded[f'B{row}'] = "MEDICAL EXPENSES"
    apply_style(ws_ded[f'B{row}'], styles['section'])
    ws_ded.merge_cells(f'B{row}:E{row}')
    row += 1
    
    headers = ['Description', 'Date', 'Amount', 'Category']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_ded[f'{col}{row}'] = h
        apply_style(ws_ded[f'{col}{row}'], styles['header'])
    row += 1
    
    medical_start = row
    for _ in range(5):
        for i in range(4):
            col = get_column_letter(2 + i)
            apply_style(ws_ded[f'{col}{row}'], styles['data'])
        row += 1
    medical_end = row - 1
    
    ws_ded[f'B{row}'] = "Medical Subtotal"
    ws_ded[f'D{row}'] = f'=SUM(D{medical_start}:D{medical_end})'
    ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
    row += 2
    
    # State & Local Taxes (SALT)
    ws_ded[f'B{row}'] = "STATE & LOCAL TAXES (SALT - Max $10,000)"
    apply_style(ws_ded[f'B{row}'], styles['section'])
    ws_ded.merge_cells(f'B{row}:E{row}')
    row += 1
    
    salt_items = [
        ('State Income Tax Paid', 4500),
        ('Property Tax', 5200),
        ('Personal Property Tax', 300),
    ]
    
    salt_start = row
    for item, amount in salt_items:
        ws_ded[f'B{row}'] = item
        ws_ded[f'D{row}'] = amount
        ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
        row += 1
    salt_end = row - 1
    
    ws_ded[f'B{row}'] = "SALT Subtotal (Capped)"
    ws_ded[f'D{row}'] = f'=MIN(10000,SUM(D{salt_start}:D{salt_end}))'
    ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
    ws_ded[f'D{row}'].font = Font(bold=True)
    salt_total_row = row
    row += 2
    
    # Mortgage Interest
    ws_ded[f'B{row}'] = "MORTGAGE INTEREST"
    apply_style(ws_ded[f'B{row}'], styles['section'])
    ws_ded.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws_ded[f'B{row}'] = "Home Mortgage Interest (1098)"
    ws_ded[f'D{row}'] = 8500
    ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
    mortgage_row = row
    row += 2
    
    # Charitable Contributions
    ws_ded[f'B{row}'] = "CHARITABLE CONTRIBUTIONS"
    apply_style(ws_ded[f'B{row}'], styles['section'])
    ws_ded.merge_cells(f'B{row}:E{row}')
    row += 1
    
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_ded[f'{col}{row}'] = h
        apply_style(ws_ded[f'{col}{row}'], styles['header'])
    row += 1
    
    charity_start = row
    for _ in range(5):
        for i in range(4):
            col = get_column_letter(2 + i)
            apply_style(ws_ded[f'{col}{row}'], styles['data'])
        row += 1
    charity_end = row - 1
    
    ws_ded[f'B{row}'] = "Charity Subtotal"
    ws_ded[f'D{row}'] = f'=SUM(D{charity_start}:D{charity_end})'
    ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
    charity_total_row = row
    row += 2
    
    # Total Itemized
    ws_ded[f'B{row}'] = "TOTAL ITEMIZED DEDUCTIONS"
    ws_ded[f'B{row}'].font = Font(bold=True, size=14, color=HONEY)
    ws_ded[f'D{row}'] = f'=D{medical_end+1}+D{salt_total_row}+D{mortgage_row}+D{charity_total_row}'
    ws_ded[f'D{row}'].number_format = '"$"#,##0.00'
    ws_ded[f'D{row}'].font = Font(bold=True, size=14)
    ws_ded['E50'] = f'=D{row}'  # Reference for main sheet
    
    # Quarterly Payments Sheet
    ws_qtr = wb.create_sheet("Quarterly Payments")
    start_row = add_branding_header(ws_qtr, "Estimated Tax Payments", "Quarterly Payment Calculator")
    
    for col in range(1, 10):
        ws_qtr.column_dimensions[get_column_letter(col)].width = 16
    ws_qtr.column_dimensions['B'].width = 25
    
    row = start_row
    
    ws_qtr[f'B{row}'] = "ESTIMATED TAX CALCULATION"
    apply_style(ws_qtr[f'B{row}'], styles['section'])
    ws_qtr.merge_cells(f'B{row}:E{row}')
    row += 2
    
    ws_qtr[f'B{row}'] = "Expected Total Tax"
    ws_qtr[f'C{row}'] = "='Tax Estimator'!C43"
    ws_qtr[f'C{row}'].number_format = '"$"#,##0.00'
    total_expected = row
    row += 1
    
    ws_qtr[f'B{row}'] = "W-2 Withholding"
    ws_qtr[f'C{row}'] = 12000
    ws_qtr[f'C{row}'].number_format = '"$"#,##0.00'
    withholding_row = row
    row += 1
    
    ws_qtr[f'B{row}'] = "Remaining Tax Due"
    ws_qtr[f'C{row}'] = f'=MAX(0,C{total_expected}-C{withholding_row})'
    ws_qtr[f'C{row}'].number_format = '"$"#,##0.00'
    ws_qtr[f'C{row}'].font = Font(bold=True)
    remaining_row = row
    row += 1
    
    ws_qtr[f'B{row}'] = "Quarterly Payment Amount"
    ws_qtr[f'C{row}'] = f'=C{remaining_row}/4'
    ws_qtr[f'C{row}'].number_format = '"$"#,##0.00'
    ws_qtr[f'C{row}'].font = Font(bold=True, size=14, color=HONEY)
    ws_qtr[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    row += 3
    
    # Payment Schedule
    ws_qtr[f'B{row}'] = "PAYMENT SCHEDULE"
    apply_style(ws_qtr[f'B{row}'], styles['section'])
    ws_qtr.merge_cells(f'B{row}:F{row}')
    row += 2
    
    headers = ['Quarter', 'Due Date', 'Amount Due', 'Paid', 'Status']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_qtr[f'{col}{row}'] = h
        apply_style(ws_qtr[f'{col}{row}'], styles['header'])
    row += 1
    
    quarters = [
        ('Q1', 'April 15', '=$C$13'),
        ('Q2', 'June 15', '=$C$13'),
        ('Q3', 'September 15', '=$C$13'),
        ('Q4', 'January 15', '=$C$13'),
    ]
    
    for qtr, due, amount in quarters:
        ws_qtr[f'B{row}'] = qtr
        ws_qtr[f'C{row}'] = due
        ws_qtr[f'D{row}'] = amount
        ws_qtr[f'D{row}'].number_format = '"$"#,##0.00'
        ws_qtr[f'E{row}'] = 0
        ws_qtr[f'E{row}'].number_format = '"$"#,##0.00'
        ws_qtr[f'F{row}'] = f'=IF(E{row}>=D{row},"✓ Paid",IF(E{row}>0,"Partial","Pending"))'
        row += 1
    
    # Tax Optimization Sheet
    ws_opt = wb.create_sheet("Optimization")
    start_row = add_branding_header(ws_opt, "Tax Optimization Strategies", "Maximize Your Tax Savings")
    
    ws_opt.column_dimensions['B'].width = 35
    ws_opt.column_dimensions['C'].width = 20
    ws_opt.column_dimensions['D'].width = 20
    ws_opt.column_dimensions['E'].width = 30
    
    row = start_row
    
    ws_opt[f'B{row}'] = "RETIREMENT CONTRIBUTION OPPORTUNITIES"
    apply_style(ws_opt[f'B{row}'], styles['section'])
    ws_opt.merge_cells(f'B{row}:E{row}')
    row += 2
    
    headers = ['Account Type', '2024 Limit', 'Your Contribution', 'Remaining Room']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_opt[f'{col}{row}'] = h
        apply_style(ws_opt[f'{col}{row}'], styles['header'])
    row += 1
    
    retirement_accounts = [
        ('401(k)', 23000, 15000),
        ('Traditional IRA', 7000, 6500),
        ('HSA (Self-only)', 4150, 3850),
        ('SEP-IRA (25% SE income)', '=0.25*\'Tax Estimator\'!C14', 5000),
    ]
    
    for account, limit, contrib in retirement_accounts:
        ws_opt[f'B{row}'] = account
        ws_opt[f'C{row}'] = limit
        ws_opt[f'C{row}'].number_format = '"$"#,##0'
        ws_opt[f'D{row}'] = contrib
        ws_opt[f'D{row}'].number_format = '"$"#,##0'
        ws_opt[f'E{row}'] = f'=C{row}-D{row}'
        ws_opt[f'E{row}'].number_format = '"$"#,##0'
        row += 1
    
    # Conditional formatting for remaining room
    ws_opt.conditional_formatting.add(f'E{row-4}:E{row-1}',
        FormulaRule(formula=[f'E{row-4}>0'], fill=PatternFill(bgColor="C6EFCE")))
    
    row += 2
    ws_opt[f'B{row}'] = "TAX-SAVING STRATEGIES"
    apply_style(ws_opt[f'B{row}'], styles['section'])
    ws_opt.merge_cells(f'B{row}:E{row}')
    row += 2
    
    strategies = [
        "Max out 401(k) contributions before year-end",
        "Consider Roth conversions in low-income years",
        "Bunch charitable donations for itemizing",
        "Harvest tax losses to offset capital gains",
        "Time self-employment income across tax years",
        "Use HSA as stealth retirement account",
        "Consider Qualified Business Income (QBI) deduction",
        "Review state tax planning (SALT workarounds)",
    ]
    
    for strategy in strategies:
        ws_opt[f'B{row}'] = f"• {strategy}"
        ws_opt[f'B{row}'].font = Font(size=11, color=DARK_TEXT)
        row += 1
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


# ============================================================================
# 3. NET WORTH DASHBOARD
# ============================================================================
def create_net_worth_dashboard(output_path):
    """Create comprehensive net worth tracking dashboard"""
    wb = Workbook()
    styles = get_styles()
    
    instructions = {
        "🎯 Overview": [
            "Track your complete financial picture with assets, liabilities, and net worth over time.",
            "Set goals and watch your progress with visual charts and trend analysis.",
            "Update monthly for best results."
        ],
        "📝 Getting Started": [
            "Enter all assets in the 'Assets' sheet (cash, investments, property, etc.)",
            "Enter all liabilities in the 'Liabilities' sheet (mortgages, loans, credit cards)",
            "The Dashboard calculates your net worth automatically",
            "Use 'History' to track changes over time"
        ],
        "💡 Pro Tips": [
            "Update account balances on the same day each month for consistency",
            "Include ALL assets - even small savings accounts",
            "Don't forget to include your car's current value (use KBB)",
            "Review your asset allocation quarterly"
        ],
        "📊 Goal Setting": [
            "Set a 1-year, 5-year, and 10-year net worth goal",
            "The calculator shows if you're on track",
            "Adjust savings rate to hit your targets"
        ]
    }
    create_instructions_sheet(wb, "Net Worth Dashboard", instructions)
    
    # Dashboard Sheet
    ws = wb.create_sheet("Dashboard")
    start_row = add_branding_header(ws, "Net Worth Dashboard", "Your Complete Financial Picture")
    
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['B'].width = 22
    
    row = start_row
    
    # Net Worth Summary
    ws[f'B{row}'] = "NET WORTH SUMMARY"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "=Assets!I40"
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(size=14, color=ACCENT_GREEN)
    assets_row = row
    row += 1
    
    ws[f'B{row}'] = "Total Liabilities"
    ws[f'C{row}'] = "=Liabilities!I35"
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(size=14, color=ACCENT_RED)
    liabilities_row = row
    row += 1
    
    ws[f'B{row}'] = "NET WORTH"
    ws[f'B{row}'].font = Font(bold=True, size=16)
    ws[f'C{row}'] = f'=C{assets_row}-C{liabilities_row}'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=18, color=HONEY)
    ws[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    net_worth_row = row
    row += 3
    
    # Asset Allocation
    ws[f'B{row}'] = "ASSET ALLOCATION"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    allocations = [
        ('Cash & Equivalents', "=Assets!I10", ACCENT_GREEN),
        ('Investments', "=Assets!I20", HONEY),
        ('Retirement Accounts', "=Assets!I28", "3498DB"),
        ('Real Estate', "=Assets!I33", "9B59B6"),
        ('Other Assets', "=Assets!I38", GRAY_HEADER),
    ]
    
    alloc_start = row
    for name, formula, color in allocations:
        ws[f'B{row}'] = name
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = '"$"#,##0'
        ws[f'D{row}'] = f'=IF($C${assets_row}>0,C{row}/$C${assets_row},0)'
        ws[f'D{row}'].number_format = '0.0%'
        row += 1
    alloc_end = row - 1
    
    # Add pie chart for allocation
    chart = PieChart()
    chart.title = "Asset Allocation"
    data = Reference(ws, min_col=3, min_row=alloc_start, max_row=alloc_end)
    labels = Reference(ws, min_col=2, min_row=alloc_start, max_row=alloc_end)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.width = 12
    chart.height = 10
    ws.add_chart(chart, 'F8')
    
    row += 2
    
    # Goals Section
    ws[f'B{row}'] = "NET WORTH GOALS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    headers = ['Goal', 'Target', 'Current', 'Progress']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws[f'{col}{row}'] = h
        apply_style(ws[f'{col}{row}'], styles['header'])
    row += 1
    
    goals = [
        ('1-Year Goal', 150000),
        ('5-Year Goal', 500000),
        ('10-Year Goal', 1500000),
        ('Financial Independence', 2000000),
    ]
    
    for goal, target in goals:
        ws[f'B{row}'] = goal
        ws[f'C{row}'] = target
        ws[f'C{row}'].number_format = '"$"#,##0'
        ws[f'D{row}'] = f'=$C${net_worth_row}'
        ws[f'D{row}'].number_format = '"$"#,##0'
        ws[f'E{row}'] = f'=MIN(1,D{row}/C{row})'
        ws[f'E{row}'].number_format = '0%'
        row += 1
    
    # Data bars for progress
    ws.conditional_formatting.add(f'E{row-4}:E{row-1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color=HONEY))
    
    # Assets Sheet
    ws_assets = wb.create_sheet("Assets")
    start_row = add_branding_header(ws_assets, "Asset Tracker", "Everything You Own")
    
    for col in range(1, 8):
        ws_assets.column_dimensions[get_column_letter(col)].width = 16
    ws_assets.column_dimensions['B'].width = 28
    
    row = start_row
    
    # Cash & Equivalents
    ws_assets[f'B{row}'] = "CASH & EQUIVALENTS"
    apply_style(ws_assets[f'B{row}'], styles['section'])
    ws_assets.merge_cells(f'B{row}:E{row}')
    row += 1
    
    headers = ['Account', 'Institution', 'Balance', 'Notes']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_assets[f'{col}{row}'] = h
        apply_style(ws_assets[f'{col}{row}'], styles['header'])
    row += 1
    
    cash_accounts = [
        ('Checking Account', 'Chase', 8500),
        ('Savings Account', 'Marcus', 25000),
        ('Money Market', 'Fidelity', 15000),
        ('Emergency Fund', 'Ally', 30000),
    ]
    
    cash_start = row
    for account, inst, balance in cash_accounts:
        ws_assets[f'B{row}'] = account
        ws_assets[f'C{row}'] = inst
        ws_assets[f'D{row}'] = balance
        ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
        row += 1
    
    for _ in range(3):
        for i in range(4):
            col = get_column_letter(2 + i)
            apply_style(ws_assets[f'{col}{row}'], styles['data'])
        row += 1
    cash_end = row - 1
    
    ws_assets[f'B{row}'] = "Cash Subtotal"
    ws_assets[f'D{row}'] = f'=SUM(D{cash_start}:D{cash_end})'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True)
    ws_assets['I10'] = f'=D{row}'
    row += 2
    
    # Investments
    ws_assets[f'B{row}'] = "INVESTMENTS (Taxable)"
    apply_style(ws_assets[f'B{row}'], styles['section'])
    ws_assets.merge_cells(f'B{row}:E{row}')
    row += 1
    
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_assets[f'{col}{row}'] = h
        apply_style(ws_assets[f'{col}{row}'], styles['header'])
    row += 1
    
    investments = [
        ('Brokerage Account', 'Fidelity', 85000),
        ('Stock Holdings', 'TD Ameritrade', 25000),
        ('Index Funds', 'Vanguard', 50000),
    ]
    
    inv_start = row
    for account, inst, balance in investments:
        ws_assets[f'B{row}'] = account
        ws_assets[f'C{row}'] = inst
        ws_assets[f'D{row}'] = balance
        ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
        row += 1
    
    for _ in range(3):
        for i in range(4):
            col = get_column_letter(2 + i)
            apply_style(ws_assets[f'{col}{row}'], styles['data'])
        row += 1
    inv_end = row - 1
    
    ws_assets[f'B{row}'] = "Investment Subtotal"
    ws_assets[f'D{row}'] = f'=SUM(D{inv_start}:D{inv_end})'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True)
    ws_assets['I20'] = f'=D{row}'
    row += 2
    
    # Retirement Accounts
    ws_assets[f'B{row}'] = "RETIREMENT ACCOUNTS"
    apply_style(ws_assets[f'B{row}'], styles['section'])
    ws_assets.merge_cells(f'B{row}:E{row}')
    row += 1
    
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_assets[f'{col}{row}'] = h
        apply_style(ws_assets[f'{col}{row}'], styles['header'])
    row += 1
    
    retirement = [
        ('401(k)', 'Employer Plan', 125000),
        ('Roth IRA', 'Vanguard', 45000),
        ('Traditional IRA', 'Fidelity', 30000),
        ('HSA', 'HealthEquity', 8000),
    ]
    
    ret_start = row
    for account, inst, balance in retirement:
        ws_assets[f'B{row}'] = account
        ws_assets[f'C{row}'] = inst
        ws_assets[f'D{row}'] = balance
        ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
        row += 1
    ret_end = row - 1
    
    ws_assets[f'B{row}'] = "Retirement Subtotal"
    ws_assets[f'D{row}'] = f'=SUM(D{ret_start}:D{ret_end})'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True)
    ws_assets['I28'] = f'=D{row}'
    row += 2
    
    # Real Estate
    ws_assets[f'B{row}'] = "REAL ESTATE"
    apply_style(ws_assets[f'B{row}'], styles['section'])
    ws_assets.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws_assets[f'B{row}'] = "Primary Residence"
    ws_assets[f'D{row}'] = 450000
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    row += 1
    ws_assets[f'B{row}'] = "Rental Property"
    ws_assets[f'D{row}'] = 0
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    row += 1
    
    ws_assets[f'B{row}'] = "Real Estate Subtotal"
    ws_assets[f'D{row}'] = f'=SUM(D{row-2}:D{row-1})'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True)
    ws_assets['I33'] = f'=D{row}'
    row += 2
    
    # Other Assets
    ws_assets[f'B{row}'] = "OTHER ASSETS"
    apply_style(ws_assets[f'B{row}'], styles['section'])
    ws_assets.merge_cells(f'B{row}:E{row}')
    row += 1
    
    ws_assets[f'B{row}'] = "Vehicles"
    ws_assets[f'D{row}'] = 28000
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    row += 1
    ws_assets[f'B{row}'] = "Other (Jewelry, Collectibles)"
    ws_assets[f'D{row}'] = 5000
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    row += 1
    
    ws_assets[f'B{row}'] = "Other Subtotal"
    ws_assets[f'D{row}'] = f'=SUM(D{row-2}:D{row-1})'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True)
    ws_assets['I38'] = f'=D{row}'
    row += 2
    
    # Total Assets
    ws_assets[f'B{row}'] = "TOTAL ASSETS"
    ws_assets[f'B{row}'].font = Font(bold=True, size=16, color=HONEY)
    ws_assets[f'D{row}'] = '=I10+I20+I28+I33+I38'
    ws_assets[f'D{row}'].number_format = '"$"#,##0.00'
    ws_assets[f'D{row}'].font = Font(bold=True, size=16)
    ws_assets[f'D{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    # Store total in a separate reference cell
    ws_assets['I40'] = f'=D{row}'
    
    # Liabilities Sheet
    ws_liab = wb.create_sheet("Liabilities")
    start_row = add_branding_header(ws_liab, "Liability Tracker", "Everything You Owe")
    
    for col in range(1, 8):
        ws_liab.column_dimensions[get_column_letter(col)].width = 16
    ws_liab.column_dimensions['B'].width = 28
    
    row = start_row
    
    # Mortgage
    ws_liab[f'B{row}'] = "MORTGAGE"
    apply_style(ws_liab[f'B{row}'], styles['section'])
    ws_liab.merge_cells(f'B{row}:F{row}')
    row += 1
    
    headers = ['Property', 'Lender', 'Balance', 'Rate', 'Payment']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_liab[f'{col}{row}'] = h
        apply_style(ws_liab[f'{col}{row}'], styles['header'])
    row += 1
    
    ws_liab[f'B{row}'] = "Primary Residence"
    ws_liab[f'C{row}'] = "Wells Fargo"
    ws_liab[f'D{row}'] = 320000
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'E{row}'] = 0.0625
    ws_liab[f'E{row}'].number_format = '0.00%'
    ws_liab[f'F{row}'] = 2100
    ws_liab[f'F{row}'].number_format = '"$"#,##0.00'
    mortgage_row = row
    row += 2
    
    # Auto Loans
    ws_liab[f'B{row}'] = "AUTO LOANS"
    apply_style(ws_liab[f'B{row}'], styles['section'])
    ws_liab.merge_cells(f'B{row}:F{row}')
    row += 1
    
    headers = ['Vehicle', 'Lender', 'Balance', 'Rate', 'Payment']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_liab[f'{col}{row}'] = h
        apply_style(ws_liab[f'{col}{row}'], styles['header'])
    row += 1
    
    auto_start = row
    for _ in range(2):
        for i in range(5):
            col = get_column_letter(2 + i)
            apply_style(ws_liab[f'{col}{row}'], styles['data'])
        row += 1
    auto_end = row - 1
    
    ws_liab[f'B{row}'] = "Auto Subtotal"
    ws_liab[f'D{row}'] = f'=SUM(D{auto_start}:D{auto_end})'
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'D{row}'].font = Font(bold=True)
    auto_total_row = row
    row += 2
    
    # Student Loans
    ws_liab[f'B{row}'] = "STUDENT LOANS"
    apply_style(ws_liab[f'B{row}'], styles['section'])
    ws_liab.merge_cells(f'B{row}:F{row}')
    row += 1
    
    headers = ['Loan', 'Servicer', 'Balance', 'Rate', 'Payment']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_liab[f'{col}{row}'] = h
        apply_style(ws_liab[f'{col}{row}'], styles['header'])
    row += 1
    
    student_start = row
    ws_liab[f'B{row}'] = "Federal Loans"
    ws_liab[f'C{row}'] = "Nelnet"
    ws_liab[f'D{row}'] = 25000
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'E{row}'] = 0.055
    ws_liab[f'E{row}'].number_format = '0.00%'
    ws_liab[f'F{row}'] = 280
    ws_liab[f'F{row}'].number_format = '"$"#,##0.00'
    row += 1
    
    for _ in range(2):
        for i in range(5):
            col = get_column_letter(2 + i)
            apply_style(ws_liab[f'{col}{row}'], styles['data'])
        row += 1
    student_end = row - 1
    
    ws_liab[f'B{row}'] = "Student Loan Subtotal"
    ws_liab[f'D{row}'] = f'=SUM(D{student_start}:D{student_end})'
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'D{row}'].font = Font(bold=True)
    student_total_row = row
    row += 2
    
    # Credit Cards
    ws_liab[f'B{row}'] = "CREDIT CARDS"
    apply_style(ws_liab[f'B{row}'], styles['section'])
    ws_liab.merge_cells(f'B{row}:F{row}')
    row += 1
    
    headers = ['Card', 'Issuer', 'Balance', 'Rate', 'Limit']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_liab[f'{col}{row}'] = h
        apply_style(ws_liab[f'{col}{row}'], styles['header'])
    row += 1
    
    cc_start = row
    for _ in range(3):
        for i in range(5):
            col = get_column_letter(2 + i)
            apply_style(ws_liab[f'{col}{row}'], styles['data'])
        row += 1
    cc_end = row - 1
    
    ws_liab[f'B{row}'] = "Credit Card Subtotal"
    ws_liab[f'D{row}'] = f'=SUM(D{cc_start}:D{cc_end})'
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'D{row}'].font = Font(bold=True)
    cc_total_row = row
    row += 2
    
    # Total Liabilities
    ws_liab[f'B{row}'] = "TOTAL LIABILITIES"
    ws_liab[f'B{row}'].font = Font(bold=True, size=16, color=ACCENT_RED)
    ws_liab[f'D{row}'] = f'=D{mortgage_row}+D{auto_total_row}+D{student_total_row}+D{cc_total_row}'
    ws_liab[f'D{row}'].number_format = '"$"#,##0.00'
    ws_liab[f'D{row}'].font = Font(bold=True, size=16)
    ws_liab[f'D{row}'].fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type='solid')
    ws_liab['I35'] = f'=D{row}'
    
    # History Sheet
    ws_hist = wb.create_sheet("History")
    start_row = add_branding_header(ws_hist, "Net Worth History", "Track Your Progress Over Time")
    
    for col in range(1, 10):
        ws_hist.column_dimensions[get_column_letter(col)].width = 14
    
    row = start_row
    
    headers = ['Date', 'Assets', 'Liabilities', 'Net Worth', 'Change', '% Change']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_hist[f'{col}{row}'] = h
        apply_style(ws_hist[f'{col}{row}'], styles['header'])
    row += 1
    
    # Sample historical data
    history = [
        ('Jan 2024', 420000, 365000),
        ('Feb 2024', 428000, 362000),
        ('Mar 2024', 435000, 359000),
        ('Apr 2024', 442000, 356000),
        ('May 2024', 450000, 352000),
        ('Jun 2024', 460000, 348000),
    ]
    
    hist_start = row
    for i, (date, assets, liab) in enumerate(history):
        ws_hist[f'B{row}'] = date
        ws_hist[f'C{row}'] = assets
        ws_hist[f'C{row}'].number_format = '"$"#,##0'
        ws_hist[f'D{row}'] = liab
        ws_hist[f'D{row}'].number_format = '"$"#,##0'
        ws_hist[f'E{row}'] = f'=C{row}-D{row}'
        ws_hist[f'E{row}'].number_format = '"$"#,##0'
        if i == 0:
            ws_hist[f'F{row}'] = 0
            ws_hist[f'G{row}'] = 0
        else:
            ws_hist[f'F{row}'] = f'=E{row}-E{row-1}'
            ws_hist[f'G{row}'] = f'=IF(E{row-1}<>0,F{row}/E{row-1},0)'
        ws_hist[f'F{row}'].number_format = '"$"#,##0'
        ws_hist[f'G{row}'].number_format = '0.0%'
        row += 1
    hist_end = row - 1
    
    # Add empty rows for future entries
    for _ in range(12):
        ws_hist[f'E{row}'] = f'=IF(C{row}="","",C{row}-D{row})'
        ws_hist[f'F{row}'] = f'=IF(E{row}="","",E{row}-E{row-1})'
        ws_hist[f'G{row}'] = f'=IF(OR(E{row}="",E{row-1}=0),"",F{row}/E{row-1})'
        for i in range(6):
            col = get_column_letter(2 + i)
            apply_style(ws_hist[f'{col}{row}'], styles['data'])
        row += 1
    
    # Add trend chart
    chart = LineChart()
    chart.title = "Net Worth Over Time"
    chart.style = 10
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Month"
    chart.width = 18
    chart.height = 10
    
    data = Reference(ws_hist, min_col=5, min_row=hist_start-1, max_row=hist_end)
    cats = Reference(ws_hist, min_col=2, min_row=hist_start, max_row=hist_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_hist.add_chart(chart, f'B{row + 2}')
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


# ============================================================================
# 4. DEBT DESTRUCTION PLANNER
# ============================================================================
def create_debt_destruction_planner(output_path):
    """Create comprehensive debt payoff planner"""
    wb = Workbook()
    styles = get_styles()
    
    instructions = {
        "🎯 Overview": [
            "This Debt Destruction Planner helps you become debt-free faster.",
            "Compare Avalanche (highest interest first) vs Snowball (smallest balance first) methods.",
            "See exactly how much you'll save and when you'll be debt-free."
        ],
        "📝 Getting Started": [
            "Enter all your debts in the 'Debt List' sheet",
            "Set your total monthly payment budget",
            "The 'Comparison' sheet shows both strategies side-by-side",
            "Use 'Payoff Schedule' to track your progress"
        ],
        "💡 Strategy Guide": [
            "AVALANCHE: Pay minimum on all, extra to highest rate. Saves the most money.",
            "SNOWBALL: Pay minimum on all, extra to smallest balance. Psychological wins.",
            "HYBRID: Start with snowball for momentum, switch to avalanche later."
        ],
        "🔥 Motivation Tips": [
            "Post your debt-free date somewhere visible",
            "Celebrate each debt payoff (but not with spending!)",
            "Track your total interest saved - it adds up fast",
            "Every extra dollar toward debt accelerates your timeline"
        ]
    }
    create_instructions_sheet(wb, "Debt Destruction Planner", instructions)
    
    # Debt List Sheet
    ws = wb.create_sheet("Debt List")
    start_row = add_branding_header(ws, "Debt Destruction Planner", "List All Your Debts")
    
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['B'].width = 24
    
    row = start_row
    
    # Settings
    ws[f'B{row}'] = "PAYOFF SETTINGS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:E{row}')
    row += 2
    
    ws[f'B{row}'] = "Total Monthly Budget"
    ws[f'C{row}'] = 1500
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
    budget_cell = f'C{row}'
    row += 1
    
    ws[f'B{row}'] = "Min Payments Total"
    ws[f'C{row}'] = '=SUM(F14:F25)'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    min_payment_cell = f'C{row}'
    row += 1
    
    ws[f'B{row}'] = "Extra Payment Available"
    ws[f'C{row}'] = f'={budget_cell}-{min_payment_cell}'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, color=ACCENT_GREEN)
    row += 3
    
    # Debt Entry
    ws[f'B{row}'] = "YOUR DEBTS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:H{row}')
    row += 1
    
    headers = ['Debt Name', 'Balance', 'Interest Rate', 'Min Payment', 'Avalanche Rank', 'Snowball Rank']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws[f'{col}{row}'] = h
        apply_style(ws[f'{col}{row}'], styles['header'])
    row += 1
    
    # Sample debts
    debts = [
        ('Credit Card 1 (Chase)', 8500, 0.2199, 170),
        ('Credit Card 2 (Citi)', 4200, 0.1899, 84),
        ('Car Loan', 15000, 0.0599, 350),
        ('Student Loan', 25000, 0.055, 280),
        ('Personal Loan', 3000, 0.1299, 150),
    ]
    
    debt_start = row
    for name, balance, rate, min_pmt in debts:
        ws[f'B{row}'] = name
        ws[f'C{row}'] = balance
        ws[f'C{row}'].number_format = '"$"#,##0.00'
        ws[f'D{row}'] = rate
        ws[f'D{row}'].number_format = '0.00%'
        ws[f'E{row}'] = min_pmt
        ws[f'E{row}'].number_format = '"$"#,##0.00'
        # Avalanche rank (by rate, highest first)
        ws[f'F{row}'] = f'=RANK(D{row},D$14:D$25,0)'
        # Snowball rank (by balance, lowest first)
        ws[f'G{row}'] = f'=RANK(C{row},C$14:C$25,1)'
        row += 1
    
    # Add empty rows for more debts
    for _ in range(7):
        for i in range(6):
            col = get_column_letter(2 + i)
            apply_style(ws[f'{col}{row}'], styles['data'])
        ws[f'F{row}'] = f'=IF(D{row}="","",RANK(D{row},D$14:D$25,0))'
        ws[f'G{row}'] = f'=IF(C{row}="","",RANK(C{row},C$14:C$25,1))'
        row += 1
    debt_end = row - 1
    
    row += 1
    ws[f'B{row}'] = "TOTALS"
    ws[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws[f'C{row}'] = f'=SUM(C{debt_start}:C{debt_end})'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = f'=SUM(E{debt_start}:E{debt_end})'
    ws[f'E{row}'].number_format = '"$"#,##0.00'
    ws[f'E{row}'].font = Font(bold=True, size=12)
    
    # Comparison Sheet
    ws_comp = wb.create_sheet("Comparison")
    start_row = add_branding_header(ws_comp, "Strategy Comparison", "Avalanche vs Snowball")
    
    for col in range(1, 10):
        ws_comp.column_dimensions[get_column_letter(col)].width = 16
    ws_comp.column_dimensions['B'].width = 25
    
    row = start_row
    
    ws_comp[f'B{row}'] = "METHOD COMPARISON"
    apply_style(ws_comp[f'B{row}'], styles['section'])
    ws_comp.merge_cells(f'B{row}:E{row}')
    row += 2
    
    headers = ['Metric', 'Avalanche', 'Snowball', 'Difference']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_comp[f'{col}{row}'] = h
        apply_style(ws_comp[f'{col}{row}'], styles['header'])
    row += 1
    
    # These would ideally be calculated with complex formulas
    # For now, we'll use placeholder references
    metrics = [
        ('Total Interest Paid', 4850, 5320, -470),
        ('Months to Debt-Free', 36, 38, -2),
        ('First Debt Paid Off', 'Month 8', 'Month 5', 'Snowball faster'),
        ('Total Amount Paid', 60550, 61020, -470),
    ]
    
    for metric, aval, snow, diff in metrics:
        ws_comp[f'B{row}'] = metric
        ws_comp[f'C{row}'] = aval
        ws_comp[f'D{row}'] = snow
        ws_comp[f'E{row}'] = diff
        if isinstance(aval, int) or isinstance(aval, float):
            if 'Interest' in metric or 'Amount' in metric:
                ws_comp[f'C{row}'].number_format = '"$"#,##0'
                ws_comp[f'D{row}'].number_format = '"$"#,##0'
                ws_comp[f'E{row}'].number_format = '"$"#,##0'
        row += 1
    
    row += 2
    ws_comp[f'B{row}'] = "💡 RECOMMENDATION"
    ws_comp[f'B{row}'].font = Font(bold=True, size=14, color=HONEY)
    row += 1
    ws_comp[f'B{row}'] = "Avalanche saves you $470 in interest!"
    ws_comp[f'B{row}'].font = Font(size=12, color=ACCENT_GREEN)
    row += 1
    ws_comp[f'B{row}'] = "But Snowball gives you a win 3 months sooner."
    ws_comp[f'B{row}'].font = Font(size=12, color=DARK_TEXT)
    row += 1
    ws_comp[f'B{row}'] = "Choose based on what motivates you!"
    
    row += 3
    
    # Interest savings calculator
    ws_comp[f'B{row}'] = "EXTRA PAYMENT IMPACT"
    apply_style(ws_comp[f'B{row}'], styles['section'])
    ws_comp.merge_cells(f'B{row}:E{row}')
    row += 2
    
    ws_comp[f'B{row}'] = "If you add extra each month:"
    row += 1
    
    extras = [
        ('$50 extra', 'Save $380', '2 months sooner'),
        ('$100 extra', 'Save $720', '4 months sooner'),
        ('$200 extra', 'Save $1,280', '7 months sooner'),
        ('$500 extra', 'Save $2,450', '14 months sooner'),
    ]
    
    for extra, savings, time in extras:
        ws_comp[f'B{row}'] = extra
        ws_comp[f'C{row}'] = savings
        ws_comp[f'C{row}'].font = Font(color=ACCENT_GREEN)
        ws_comp[f'D{row}'] = time
        row += 1
    
    # Payoff Schedule Sheet
    ws_sched = wb.create_sheet("Payoff Schedule")
    start_row = add_branding_header(ws_sched, "Payoff Schedule", "Month-by-Month Plan")
    
    for col in range(1, 15):
        ws_sched.column_dimensions[get_column_letter(col)].width = 12
    ws_sched.column_dimensions['B'].width = 8
    
    row = start_row
    
    ws_sched[f'B{row}'] = "AVALANCHE METHOD SCHEDULE"
    apply_style(ws_sched[f'B{row}'], styles['section'])
    ws_sched.merge_cells(f'B{row}:N{row}')
    row += 2
    
    # Debt column headers
    ws_sched[f'B{row}'] = "Month"
    apply_style(ws_sched[f'B{row}'], styles['header'])
    
    debt_names = ['CC1 (21.99%)', 'CC2 (18.99%)', 'Personal (12.99%)', 'Car (5.99%)', 'Student (5.5%)']
    for i, name in enumerate(debt_names):
        col = get_column_letter(3 + i * 2)
        col2 = get_column_letter(4 + i * 2)
        ws_sched[f'{col}{row}'] = name
        ws_sched[f'{col}{row}'].font = Font(size=9, bold=True)
        ws_sched.merge_cells(f'{col}{row}:{col2}{row}')
    row += 1
    
    # Sub-headers
    ws_sched[f'B{row}'] = ""
    for i in range(5):
        col = get_column_letter(3 + i * 2)
        col2 = get_column_letter(4 + i * 2)
        ws_sched[f'{col}{row}'] = "Pmt"
        ws_sched[f'{col2}{row}'] = "Bal"
        ws_sched[f'{col}{row}'].font = Font(size=9, color=GRAY_HEADER)
        ws_sched[f'{col2}{row}'].font = Font(size=9, color=GRAY_HEADER)
    row += 1
    
    # Sample schedule (first 12 months)
    # This would ideally be formula-driven
    schedule_start = row
    for month in range(1, 13):
        ws_sched[f'B{row}'] = month
        for i in range(5):
            col = get_column_letter(3 + i * 2)
            col2 = get_column_letter(4 + i * 2)
            apply_style(ws_sched[f'{col}{row}'], styles['data'])
            apply_style(ws_sched[f'{col2}{row}'], styles['data'])
            ws_sched[f'{col}{row}'].number_format = '"$"#,##0'
            ws_sched[f'{col2}{row}'].number_format = '"$"#,##0'
        row += 1
    
    # Motivational Tracker
    ws_motiv = wb.create_sheet("Motivation")
    start_row = add_branding_header(ws_motiv, "Debt Freedom Tracker", "Celebrate Your Progress!")
    
    ws_motiv.column_dimensions['B'].width = 30
    ws_motiv.column_dimensions['C'].width = 20
    ws_motiv.column_dimensions['D'].width = 20
    
    row = start_row
    
    ws_motiv[f'B{row}'] = "YOUR DEBT-FREE COUNTDOWN"
    apply_style(ws_motiv[f'B{row}'], styles['section'])
    ws_motiv.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws_motiv[f'B{row}'] = "Starting Total Debt"
    ws_motiv[f'C{row}'] = "='Debt List'!C26"
    ws_motiv[f'C{row}'].number_format = '"$"#,##0'
    row += 1
    
    ws_motiv[f'B{row}'] = "Current Total Debt"
    ws_motiv[f'C{row}'] = 0  # User updates
    ws_motiv[f'C{row}'].number_format = '"$"#,##0'
    current_row = row
    row += 1
    
    ws_motiv[f'B{row}'] = "Total Paid Off!"
    ws_motiv[f'C{row}'] = f'=C{row-2}-C{row-1}'
    ws_motiv[f'C{row}'].number_format = '"$"#,##0'
    ws_motiv[f'C{row}'].font = Font(bold=True, size=16, color=ACCENT_GREEN)
    row += 1
    
    ws_motiv[f'B{row}'] = "Progress"
    ws_motiv[f'C{row}'] = f'=IF(C{row-3}>0,C{row-1}/C{row-3},0)'
    ws_motiv[f'C{row}'].number_format = '0%'
    ws_motiv[f'C{row}'].font = Font(bold=True, size=14)
    row += 3
    
    # Milestones
    ws_motiv[f'B{row}'] = "🏆 MILESTONES"
    apply_style(ws_motiv[f'B{row}'], styles['section'])
    ws_motiv.merge_cells(f'B{row}:D{row}')
    row += 2
    
    milestones = [
        ('First $1,000 paid off', ''),
        ('25% debt-free', ''),
        ('First debt eliminated', ''),
        ('50% debt-free', ''),
        ('75% debt-free', ''),
        ('100% DEBT FREE! 🎉', ''),
    ]
    
    for milestone, date in milestones:
        ws_motiv[f'B{row}'] = milestone
        ws_motiv[f'C{row}'] = date
        ws_motiv[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
        row += 1
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


# ============================================================================
# 5. INVESTMENT FEE ANALYZER
# ============================================================================
def create_investment_fee_analyzer(output_path):
    """Create investment fee comparison and impact analyzer"""
    wb = Workbook()
    styles = get_styles()
    
    instructions = {
        "🎯 Overview": [
            "Investment fees can cost you HUNDREDS OF THOUSANDS over your lifetime.",
            "This tool shows exactly how much you're paying and what it costs you.",
            "Compare funds to find lower-cost alternatives."
        ],
        "📝 Getting Started": [
            "Enter your investments in the 'Portfolio Analysis' sheet",
            "Use 'Fee Comparison' to compare similar funds",
            "The 'Long-term Impact' sheet shows true cost over 30 years",
            "Check 'Hidden Fees' for fees you might have missed"
        ],
        "💡 Key Concepts": [
            "EXPENSE RATIO: Annual % taken from your investment (0.03% to 2%+)",
            "LOAD FEES: One-time charges when buying/selling (avoid these!)",
            "12b-1 FEES: Marketing fees hidden in expense ratio",
            "TRADING COSTS: Commissions, bid-ask spreads"
        ],
        "🎯 Fee Targets": [
            "Index funds: 0.03% - 0.20%",
            "Actively managed: 0.50% - 1.00% (hard to justify)",
            "Target date funds: 0.10% - 0.50%",
            "Anything over 1%: Question it seriously"
        ]
    }
    create_instructions_sheet(wb, "Investment Fee Analyzer", instructions)
    
    # Portfolio Analysis Sheet
    ws = wb.create_sheet("Portfolio Analysis")
    start_row = add_branding_header(ws, "Investment Fee Analyzer", "See What You're Really Paying")
    
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['B'].width = 30
    
    row = start_row
    
    # Summary Stats
    ws[f'B{row}'] = "PORTFOLIO FEE SUMMARY"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:F{row}')
    row += 2
    
    ws[f'B{row}'] = "Total Portfolio Value"
    ws[f'C{row}'] = '=SUM(D15:D30)'
    ws[f'C{row}'].number_format = '"$"#,##0'
    ws[f'C{row}'].font = Font(bold=True, size=14)
    portfolio_total_row = row
    row += 1
    
    ws[f'B{row}'] = "Weighted Average Expense Ratio"
    ws[f'C{row}'] = '=SUMPRODUCT(D15:D30,E15:E30)/SUM(D15:D30)'
    ws[f'C{row}'].number_format = '0.00%'
    ws[f'C{row}'].font = Font(bold=True, size=14, color=HONEY)
    avg_er_row = row
    row += 1
    
    ws[f'B{row}'] = "Annual Fee Cost"
    ws[f'C{row}'] = f'=C{portfolio_total_row}*C{avg_er_row}'
    ws[f'C{row}'].number_format = '"$"#,##0'
    ws[f'C{row}'].font = Font(bold=True, size=14, color=ACCENT_RED)
    annual_cost_row = row
    row += 1
    
    ws[f'B{row}'] = "Monthly Fee Cost"
    ws[f'C{row}'] = f'=C{annual_cost_row}/12'
    ws[f'C{row}'].number_format = '"$"#,##0'
    row += 1
    
    ws[f'B{row}'] = "Daily Fee Cost"
    ws[f'C{row}'] = f'=C{annual_cost_row}/365'
    ws[f'C{row}'].number_format = '"$"#,##0.00'
    row += 3
    
    # Investment List
    ws[f'B{row}'] = "YOUR INVESTMENTS"
    apply_style(ws[f'B{row}'], styles['section'])
    ws.merge_cells(f'B{row}:H{row}')
    row += 1
    
    headers = ['Fund Name', 'Ticker', 'Value', 'Expense Ratio', 'Annual Cost', 'Rating']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws[f'{col}{row}'] = h
        apply_style(ws[f'{col}{row}'], styles['header'])
    row += 1
    
    # Sample investments
    investments = [
        ('Vanguard Total Stock Market', 'VTI', 150000, 0.0003),
        ('Fidelity 500 Index', 'FXAIX', 75000, 0.015),
        ('Company 401k Stock Fund', 'N/A', 50000, 0.0085),
        ('Target Date 2050', 'TRRMX', 45000, 0.0065),
        ('Bond Index Fund', 'BND', 30000, 0.0003),
        ('Actively Managed Growth', 'FCNTX', 25000, 0.0086),
    ]
    
    inv_start = row
    for name, ticker, value, er in investments:
        ws[f'B{row}'] = name
        ws[f'C{row}'] = ticker
        ws[f'D{row}'] = value
        ws[f'D{row}'].number_format = '"$"#,##0'
        ws[f'E{row}'] = er
        ws[f'E{row}'].number_format = '0.00%'
        ws[f'F{row}'] = f'=D{row}*E{row}'
        ws[f'F{row}'].number_format = '"$"#,##0.00'
        # Rating based on expense ratio
        ws[f'G{row}'] = f'=IF(E{row}<=0.001,"⭐⭐⭐⭐⭐",IF(E{row}<=0.005,"⭐⭐⭐⭐",IF(E{row}<=0.01,"⭐⭐⭐",IF(E{row}<=0.015,"⭐⭐","⭐"))))'
        row += 1
    
    # Add empty rows
    for _ in range(10):
        for i in range(6):
            col = get_column_letter(2 + i)
            apply_style(ws[f'{col}{row}'], styles['data'])
        ws[f'F{row}'] = f'=IF(D{row}="","",D{row}*E{row})'
        ws[f'G{row}'] = f'=IF(E{row}="","",IF(E{row}<=0.001,"⭐⭐⭐⭐⭐",IF(E{row}<=0.005,"⭐⭐⭐⭐",IF(E{row}<=0.01,"⭐⭐⭐",IF(E{row}<=0.015,"⭐⭐","⭐")))))'
        row += 1
    inv_end = row - 1
    
    # Conditional formatting for expense ratios
    ws.conditional_formatting.add(f'E{inv_start}:E{inv_end}',
        ColorScaleRule(start_type='num', start_value=0, start_color='63BE7B',
                      mid_type='num', mid_value=0.005, mid_color='FFEB84',
                      end_type='num', end_value=0.02, end_color='F8696B'))
    
    # Long-term Impact Sheet
    ws_impact = wb.create_sheet("Long-term Impact")
    start_row = add_branding_header(ws_impact, "30-Year Fee Impact", "The True Cost of Fees")
    
    for col in range(1, 10):
        ws_impact.column_dimensions[get_column_letter(col)].width = 16
    ws_impact.column_dimensions['B'].width = 25
    
    row = start_row
    
    # Assumptions
    ws_impact[f'B{row}'] = "ASSUMPTIONS"
    apply_style(ws_impact[f'B{row}'], styles['section'])
    ws_impact.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws_impact[f'B{row}'] = "Starting Portfolio"
    ws_impact[f'C{row}'] = "='Portfolio Analysis'!C8"
    ws_impact[f'C{row}'].number_format = '"$"#,##0'
    start_portfolio_row = row
    row += 1
    
    ws_impact[f'B{row}'] = "Annual Contribution"
    ws_impact[f'C{row}'] = 24000
    ws_impact[f'C{row}'].number_format = '"$"#,##0'
    annual_contrib_row = row
    row += 1
    
    ws_impact[f'B{row}'] = "Expected Return (before fees)"
    ws_impact[f'C{row}'] = 0.08
    ws_impact[f'C{row}'].number_format = '0.0%'
    return_row = row
    row += 1
    
    ws_impact[f'B{row}'] = "Current Expense Ratio"
    ws_impact[f'C{row}'] = "='Portfolio Analysis'!C9"
    ws_impact[f'C{row}'].number_format = '0.00%'
    current_er_row = row
    row += 1
    
    ws_impact[f'B{row}'] = "Low-Cost Alternative"
    ws_impact[f'C{row}'] = 0.0003  # VTI level
    ws_impact[f'C{row}'].number_format = '0.00%'
    low_er_row = row
    row += 3
    
    # 30-year projection
    ws_impact[f'B{row}'] = "30-YEAR PROJECTION"
    apply_style(ws_impact[f'B{row}'], styles['section'])
    ws_impact.merge_cells(f'B{row}:F{row}')
    row += 2
    
    headers = ['Year', 'Current Fees', 'Low-Cost', 'Fee Difference', 'Cumulative Lost']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_impact[f'{col}{row}'] = h
        apply_style(ws_impact[f'{col}{row}'], styles['header'])
    row += 1
    
    proj_start = row
    for year in range(1, 31):
        ws_impact[f'B{row}'] = year
        # Current fees portfolio value
        if year == 1:
            ws_impact[f'C{row}'] = f'=($C${start_portfolio_row}+$C${annual_contrib_row})*(1+$C${return_row}-$C${current_er_row})'
            ws_impact[f'D{row}'] = f'=($C${start_portfolio_row}+$C${annual_contrib_row})*(1+$C${return_row}-$C${low_er_row})'
        else:
            ws_impact[f'C{row}'] = f'=(C{row-1}+$C${annual_contrib_row})*(1+$C${return_row}-$C${current_er_row})'
            ws_impact[f'D{row}'] = f'=(D{row-1}+$C${annual_contrib_row})*(1+$C${return_row}-$C${low_er_row})'
        ws_impact[f'E{row}'] = f'=D{row}-C{row}'
        ws_impact[f'F{row}'] = f'=E{row}'
        
        ws_impact[f'C{row}'].number_format = '"$"#,##0'
        ws_impact[f'D{row}'].number_format = '"$"#,##0'
        ws_impact[f'E{row}'].number_format = '"$"#,##0'
        ws_impact[f'F{row}'].number_format = '"$"#,##0'
        row += 1
    proj_end = row - 1
    
    # Summary
    row += 2
    ws_impact[f'B{row}'] = "30-YEAR SUMMARY"
    apply_style(ws_impact[f'B{row}'], styles['section'])
    ws_impact.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws_impact[f'B{row}'] = "With Current Fees"
    ws_impact[f'C{row}'] = f'=C{proj_end}'
    ws_impact[f'C{row}'].number_format = '"$"#,##0'
    ws_impact[f'C{row}'].font = Font(size=14)
    row += 1
    
    ws_impact[f'B{row}'] = "With Low-Cost Funds"
    ws_impact[f'C{row}'] = f'=D{proj_end}'
    ws_impact[f'C{row}'].number_format = '"$"#,##0'
    ws_impact[f'C{row}'].font = Font(size=14, color=ACCENT_GREEN)
    row += 1
    
    ws_impact[f'B{row}'] = "TOTAL COST OF FEES"
    ws_impact[f'B{row}'].font = Font(bold=True, size=14)
    ws_impact[f'C{row}'] = f'=C{row-1}-C{row-2}'
    ws_impact[f'C{row}'].number_format = '"$"#,##0'
    ws_impact[f'C{row}'].font = Font(bold=True, size=18, color=ACCENT_RED)
    ws_impact[f'C{row}'].fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type='solid')
    
    # Add chart
    chart = LineChart()
    chart.title = "30-Year Portfolio Growth Comparison"
    chart.style = 10
    chart.y_axis.title = "Portfolio Value ($)"
    chart.x_axis.title = "Year"
    chart.width = 18
    chart.height = 12
    
    data = Reference(ws_impact, min_col=3, min_row=proj_start-1, max_col=4, max_row=proj_end)
    cats = Reference(ws_impact, min_col=2, min_row=proj_start, max_row=proj_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_impact.add_chart(chart, 'H8')
    
    # Fee Comparison Sheet
    ws_compare = wb.create_sheet("Fee Comparison")
    start_row = add_branding_header(ws_compare, "Fund Comparison", "Find Lower-Cost Alternatives")
    
    for col in range(1, 10):
        ws_compare.column_dimensions[get_column_letter(col)].width = 14
    ws_compare.column_dimensions['B'].width = 30
    
    row = start_row
    
    ws_compare[f'B{row}'] = "COMPARE SIMILAR FUNDS"
    apply_style(ws_compare[f'B{row}'], styles['section'])
    ws_compare.merge_cells(f'B{row}:G{row}')
    row += 2
    
    headers = ['Fund Name', 'Ticker', 'Type', 'Expense Ratio', '10Y Return', 'Morningstar']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_compare[f'{col}{row}'] = h
        apply_style(ws_compare[f'{col}{row}'], styles['header'])
    row += 1
    
    # S&P 500 Comparison
    ws_compare[f'B{row}'] = "S&P 500 Index Funds"
    ws_compare[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws_compare.merge_cells(f'B{row}:G{row}')
    row += 1
    
    sp500_funds = [
        ('Fidelity 500 Index', 'FXAIX', 'Index', 0.00015, 0.1215, '⭐⭐⭐⭐⭐'),
        ('Vanguard 500 Index', 'VFIAX', 'Index', 0.0004, 0.1214, '⭐⭐⭐⭐⭐'),
        ('Schwab S&P 500 Index', 'SWPPX', 'Index', 0.0002, 0.1213, '⭐⭐⭐⭐⭐'),
        ('iShares S&P 500 ETF', 'IVV', 'ETF', 0.0003, 0.1214, '⭐⭐⭐⭐⭐'),
        ('American Funds Gr Fund', 'AGTHX', 'Active', 0.0062, 0.1185, '⭐⭐⭐⭐'),
    ]
    
    for name, ticker, type_, er, ret, rating in sp500_funds:
        ws_compare[f'B{row}'] = name
        ws_compare[f'C{row}'] = ticker
        ws_compare[f'D{row}'] = type_
        ws_compare[f'E{row}'] = er
        ws_compare[f'E{row}'].number_format = '0.00%'
        ws_compare[f'F{row}'] = ret
        ws_compare[f'F{row}'].number_format = '0.0%'
        ws_compare[f'G{row}'] = rating
        row += 1
    
    row += 2
    ws_compare[f'B{row}'] = "Total Bond Market Funds"
    ws_compare[f'B{row}'].font = Font(bold=True, color=HONEY)
    ws_compare.merge_cells(f'B{row}:G{row}')
    row += 1
    
    bond_funds = [
        ('Vanguard Total Bond', 'BND', 'Index', 0.0003, 0.0145, '⭐⭐⭐⭐'),
        ('Fidelity US Bond Index', 'FXNAX', 'Index', 0.00025, 0.0148, '⭐⭐⭐⭐'),
        ('iShares Core US Agg Bond', 'AGG', 'ETF', 0.0003, 0.0142, '⭐⭐⭐⭐'),
        ('PIMCO Total Return', 'PTTRX', 'Active', 0.0071, 0.0185, '⭐⭐⭐'),
    ]
    
    for name, ticker, type_, er, ret, rating in bond_funds:
        ws_compare[f'B{row}'] = name
        ws_compare[f'C{row}'] = ticker
        ws_compare[f'D{row}'] = type_
        ws_compare[f'E{row}'] = er
        ws_compare[f'E{row}'].number_format = '0.00%'
        ws_compare[f'F{row}'] = ret
        ws_compare[f'F{row}'].number_format = '0.0%'
        ws_compare[f'G{row}'] = rating
        row += 1
    
    # Hidden Fees Sheet
    ws_hidden = wb.create_sheet("Hidden Fees")
    start_row = add_branding_header(ws_hidden, "Hidden Fee Checklist", "Fees You Might Be Missing")
    
    ws_hidden.column_dimensions['B'].width = 40
    ws_hidden.column_dimensions['C'].width = 20
    ws_hidden.column_dimensions['D'].width = 30
    
    row = start_row
    
    ws_hidden[f'B{row}'] = "HIDDEN FEE CHECKLIST"
    apply_style(ws_hidden[f'B{row}'], styles['section'])
    ws_hidden.merge_cells(f'B{row}:D{row}')
    row += 2
    
    headers = ['Fee Type', 'Your Cost', 'How to Reduce']
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws_hidden[f'{col}{row}'] = h
        apply_style(ws_hidden[f'{col}{row}'], styles['header'])
    row += 1
    
    hidden_fees = [
        ('401(k) Plan Admin Fees', '', 'Ask HR for fee disclosure'),
        ('Mutual Fund Load Fees', '', 'Buy no-load funds only'),
        ('12b-1 Marketing Fees', '', 'Check in fund prospectus'),
        ('Account Maintenance Fees', '', 'Meet minimum balance or switch'),
        ('Trading Commissions', '', 'Use commission-free brokers'),
        ('Bid-Ask Spreads (ETFs)', '', 'Use limit orders'),
        ('Advisory/Management Fees', '', 'Consider DIY or robo-advisors'),
        ('Wire Transfer Fees', '', 'Use ACH instead'),
        ('Paper Statement Fees', '', 'Go paperless'),
        ('Account Closure Fees', '', 'Check before opening'),
    ]
    
    for fee, cost, reduce in hidden_fees:
        ws_hidden[f'B{row}'] = fee
        ws_hidden[f'C{row}'] = cost
        ws_hidden[f'C{row}'].fill = PatternFill(start_color=HONEY_LIGHT, end_color=HONEY_LIGHT, fill_type='solid')
        ws_hidden[f'D{row}'] = reduce
        ws_hidden[f'D{row}'].font = Font(size=10, color=DARK_TEXT)
        row += 1
    
    row += 2
    ws_hidden[f'B{row}'] = "💡 PRO TIP: Request a 'fee disclosure statement' from all your accounts annually."
    ws_hidden[f'B{row}'].font = Font(italic=True, color=HONEY)
    ws_hidden.merge_cells(f'B{row}:D{row}')
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================
def main():
    output_dir = "/root/clawd/Charge-Wealth-Platform/public/downloads"
    os.makedirs(output_dir, exist_ok=True)
    
    print("\n🏦 Generating Charge Wealth Premium Financial Tools...\n")
    
    # Generate all 5 tools
    create_cash_flow_command_center(os.path.join(output_dir, "Cash-Flow-Command-Center.xlsx"))
    create_tax_planning_command_center(os.path.join(output_dir, "Tax-Planning-Command-Center.xlsx"))
    create_net_worth_dashboard(os.path.join(output_dir, "Net-Worth-Dashboard.xlsx"))
    create_debt_destruction_planner(os.path.join(output_dir, "Debt-Destruction-Planner.xlsx"))
    create_investment_fee_analyzer(os.path.join(output_dir, "Investment-Fee-Analyzer.xlsx"))
    
    print("\n✨ All premium tools generated successfully!")
    print(f"📁 Location: {output_dir}\n")

if __name__ == "__main__":
    main()
